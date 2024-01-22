import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from lxml import etree as ET

HEADER_ROWS = 2

def parse_xml(file):
    namespaces = {
        'xfa': 'http://www.xfa.org/schema/xfa-data/1.0/',
        'xhtml': 'http://www.w3.org/1999/xhtml'
    }

    tree = ET.parse(file)
    root = tree.getroot()

    # Debug, print structure under TeamInformation
    for team_info in root.findall('.//TeamInformation', namespaces=namespaces):
        ET.dump(team_info)  # Prints the subtree of this element

    def safe_find(xpath, namespaces=namespaces):
        nodes = root.xpath(xpath, namespaces=namespaces)
        if not nodes or nodes[0].text is None:
            print(f"Node not found or empty for XPath: {xpath} in file {file}")
            return None
        return nodes[0]

    def safe_find_all(xpath, namespaces=namespaces):
        nodes = root.xpath(xpath, namespaces=namespaces)
        if not nodes:
            print(f"No nodes found for XPath: {xpath} in file {file}")
        return nodes

    # Check if it's a team or individual
    team_or_individual_xpath = './/Team_or_individual_table/Team_or_individual_table_Row2/Team_or_individual_Radio_buttons/Radio_button_list'
    team_or_individual_node = safe_find(team_or_individual_xpath, namespaces)

    if team_or_individual_node is None:
        print(f"Could not determine if team or individual for file {file}")
        return None

    team_or_individual = team_or_individual_node.text

    if team_or_individual == '1':  # Individual
        # Define individual elements

        individual_elements = {
            'Last name': './/Individual_information_Row2/Individual_information_Last_name',
            'First name': './/Individual_information_Row2/Individual_information_First_name',
            'Email': './/Individual_information_Row3/Individual_information_Email',
            'Language': './/Individual_information_Row3/Individual_information_Language',
            'Branch/Office': './/Individual_information_Row4/Individual_information_BranchOffice',
            'Region': './/Individual_information_Row5/Individual_information_Region',
            'Category': './/Category_drop_down_list',
            'Title': './/Branch_endorsement_table_Row4/Branch_head_information_title',
            ' Last name': './/Branch_endorsement_table_Row5/Branch_head_information_last_name',
            ' First name': './/Branch_endorsement_table_Row5/Branch_head_information_first_name',
            ' Branch/Office': './/Branch_endorsement_table_Row6/Branch_head_information_branch_office',
            'Last name  ': './/Nominator_table_Row2/Nominator_last_name',
            'First name  ': './/Nominator_table_Row2/Nominator_first_name',
            'Email  ': './/Nominator_table_Row3/Nominator_email',
            'Language  ': './/Nominator_table_Row3/Nominator_language',
            'Branch/Office  ': './/Nominator_table_Row4/Nominator_branch_office',
            'Details': './/Summary_of_accomplishments_input_box',
            'Type of document 1': './/Support_type_link_table_Row3/Type_Row1',
            'Hyperlink to document 1': './/Support_type_link_table_Row3/Link_Row1',
            'Type of document 2': './/Support_type_link_table_Row4/Type_Row2',
            'Hyperlink to document 2': './/Support_type_link_table_Row4/Link_Row2',
            'Type of document 3': './/Support_type_link_table_Row5/Type_Row3',
            'Hyperlink to document 3': './/Support_type_link_table_Row5/Link_Row3'
        }

        # Extract individual elements and assign values directly (not as a list)
        individual_data = {}

        individual_data['Nomination type'] = 'Individual'

        for key, xpath in individual_elements.items():
            node = root.find(xpath, namespaces=namespaces)
            individual_data[key] = node.text.strip() if node is not None and node.text is not None else ''

        # Create DataFrame for individual data
        df = pd.DataFrame([individual_data])

        excel_column_order = [
            'Nomination type',
            'Last name',
            'First name',
            'Email',
            'Language',
            'Branch/Office',
            'Region',
            'Team name (English)',
            'Team name (French)',
            'Number of members',
            'Last name ',
            'First name ',
            'Language ',
            'Branch/Office ',
            'Email ',
            'Region ',
            'Category',
            'Last name  ',
            'First name  ',
            'Email  ',
            'Language  ',
            'Branch/Office  ',
            'Details',
            'Type of document 1',
            'Hyperlink to document 1',
            'Type of document 2',
            'Hyperlink to document 2',
            'Type of document 3',
            'Hyperlink to document 3',
            'Title',
            ' Last name',
            ' First name',
            ' Branch/Office',
        ]

        # Initialize missing columns with default values
        for column in excel_column_order:
            if column not in df.columns:
                df[column] = ''

        df = df[excel_column_order]

    elif team_or_individual == '2':  # Team
         # Define team elements
        team_elements = {
            'Team name (English)': './/Team_information/HeaderRow/EnglishTeam',
            'Team name (French)': './/Team_information/HeaderRow/FrenchTeam',
            'Number of members': './/Team_information/HeaderRow/Number_of_Members',
            'Category': './/Category_drop_down_list',
            'Title': './/Branch_endorsement_table/Branch_endorsement_table_Row4/Branch_head_information_title',
            ' Last name': './/Branch_endorsement_table/Branch_endorsement_table_Row5/Branch_head_information_last_name',
            ' First name': './/Branch_endorsement_table/Branch_endorsement_table_Row5/Branch_head_information_first_name',
            ' Branch/Office': './/Branch_endorsement_table/Branch_endorsement_table_Row6/Branch_head_information_branch_office',
            'Last name  ': './/Nominator_table/Nominator_table_Row2/Nominator_last_name',
            'First name  ': './/Nominator_table/Nominator_table_Row2/Nominator_first_name',
            'Email  ': './/Nominator_table/Nominator_table_Row3/Nominator_email',
            'Language  ': './/Nominator_table/Nominator_table_Row3/Nominator_language',
            'Branch/Office  ': './/Nominator_table/Nominator_table_Row4/Nominator_branch_office',
            'Details': './/Summary_of_accomplishments/Summary_of_accomplishments_input_box',
            'Type of document 1': './/Support_type_link_table/Support_type_link_table_Row3/Type_Row1',
            'Hyperlink to document 1': './/Support_type_link_table/Support_type_link_table_Row3/Link_Row1',
            'Type of document 2': './/Support_type_link_table/Support_type_link_table_Row4/Type_Row2',
            'Hyperlink to document 2': './/Support_type_link_table/Support_type_link_table_Row4/Link_Row2',
            'Type of document 3': './/Support_type_link_table/Support_type_link_table_Row5/Type_Row3',
            'Hyperlink to document 3': './/Support_type_link_table/Support_type_link_table_Row5/Link_Row3'
        }

        # Extract team-level data
        team_data = {}
        for key, xpath in team_elements.items():
             try:
                 node = safe_find(xpath, namespaces=namespaces)
                 team_data[key] = node.text.strip() if node is not None and node.text is not None else ''
             except Exception as e:
                 print(f"Error processing team-level key '{key}' with xpath '{xpath}' in file {file}: {e}")
                 team_data[key] = ''

    # Define team member elements
        team_member_elements = {
            'Last name ': './/TeamInformation/Team_information/InformationRow/Last_name_row',
            'First name ': './/TeamInformation/Team_information/InformationRow/First_name_row',
            'Language ': './/TeamInformation/Team_information/InformationRow/Language_row',
            'Branch/Office ': './/TeamInformation/Team_information/InformationRow/Branch_office_row',
            'Email ': './/TeamInformation/Team_information/InformationRow/Email_row',
            'Region ': './/TeamInformation/Team_information/InformationRow/Region_row',
        }

        # Initialize data dictionary with team member keys
        data = {key: [] for key in team_member_elements.keys()}

        # Extract team member elements
        for key, xpath in team_member_elements.items():
             nodes = safe_find_all(xpath, namespaces=namespaces)
             for node in nodes:
                 data[key].append(node.text.strip() if node is not None and node.text is not None else '')

        team_data['Nomination type'] = 'Team'

        # Initialize keys in data for team-level information
        for key in team_data:
            if key not in data:
                data[key] = []

        # Adding team-level data to each member row
        for i in range(len(data.get('Last name ', []))):  # Assuming at least one member exists
            for key in team_data:
                data[key].append(team_data[key])

        # Create DataFrame
        df = pd.DataFrame.from_dict(data)

        excel_column_order = [
            'Nomination type',
            'Last name',
            'First name',
            'Email',
            'Language',
            'Branch/Office',
            'Region',
            'Team name (English)',
            'Team name (French)',
            'Number of members',
            'Last name ',
            'First name ',
            'Language ',
            'Branch/Office ',
            'Email ',
            'Region ',
            'Category',
            'Last name  ',
            'First name  ',
            'Email  ',
            'Language  ',
            'Branch/Office  ',
            'Details',
            'Type of document 1',
            'Hyperlink to document 1',
            'Type of document 2',
            'Hyperlink to document 2',
            'Type of document 3',
            'Hyperlink to document 3',
            'Title',
            ' Last name',
            ' First name',
            ' Branch/Office',
        ]

        # Initialize missing columns with default values
        for column in excel_column_order:
         if column not in df.columns:
             df[column] = ''

        df = df[excel_column_order]

    df.insert(0, 'ExtraColumn', '')
    return df

# Directory containing XML files
directory = 'xml_files'

# Create a backup copy of the Excel file
shutil.copy2('output.xlsx', 'output_backup.xlsx')

# Load existing Excel file into a DataFrame
# book = load_workbook('output.xlsx')

# Read columns to verify
# df = pd.read_excel('output.xlsx', header=2)

# Print the column names
# print("Column names:")
# print(df.columns.tolist())

# Iterate over all XML files in the directory
for filename in os.listdir(directory):
    if filename.endswith('.xml'):
        try:
            # Parse XML file to DataFrame
            df = parse_xml(os.path.join(directory, filename))

            # Append DataFrame to the Excel file using append mode
            with pd.ExcelWriter('output.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Assuming 'Nominations' is the sheet name
                if 'Nominations' in writer.book.sheetnames:
                    startrow = writer.book['Nominations'].max_row
                else:
                    startrow = 0

                # Append data without writing the header if the file is not empty
                header = False if startrow > 0 else True

                # Write data to 'Nominations' sheet
                df.to_excel(writer, sheet_name='Nominations', startrow=startrow, index=False, header=header)

        except Exception as e:
            print(f"An error occurred while processing the file {filename}: {e}")