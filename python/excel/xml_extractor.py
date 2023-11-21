'''
This program extracts information from multiple XML files, then populates the corresponding columns in an existing Excel Sheet as new rows.
'''

import os
import shutil
import pandas as pd
from openpyxl import load_workbook
import xml.etree.ElementTree as ET

HEADER_ROWS = 1

# Function to parse XML file and return a pandas DataFrame
def parse_xml(file):
    # Parse XML file
    tree = ET.parse(file)
    root = tree.getroot()

    # Extract data
    last_name = root.find('.//Section_A_Row1_Last_name').text
    first_name = root.find('.//Section_A_Row1_First_name').text
    pri = root.find('.//Section_A_Row2_PRI').text
    region = root.find('.//Section_A_Row7_Region').text

    # Create DataFrame
    data = {'Last Name': [last_name], 'First Name': [first_name], 'PRI': [pri], 'Region': [region]}
    df = pd.DataFrame(data)
    
    return df

# Directory containing XML files
directory = 'xml_files'

# Create a backup copy of the Excel file
shutil.copy2('output.xlsx', 'output_backup.xlsx')

# Load existing Excel file
book = load_workbook('output.xlsx')
writer = pd.ExcelWriter('output.xlsx', engine='openpyxl') 
writer.book = book

# Specify the name of the sheet where you want to append the data
sheet_name = 'Nominations'

# Get the last non-empty row in the specified sheet
last_non_empty_row = max((r[0].row for r in writer.sheets[sheet_name].iter_rows() if any(cell.value is not None for cell in r)))

# Get the column numbers for the columns in the DataFrame
column_numbers = {cell.value: i+1 for i, cell in enumerate(next(writer.sheets[sheet_name].iter_rows(min_row=1+HEADER_ROWS, max_row=1+HEADER_ROWS)))}
# Skip the HEADER_ROWS when looking for column names

# Iterate over all XML files in the directory
for filename in os.listdir(directory):
    if filename.endswith('.xml'):
        try:
            # Parse XML file to DataFrame
            df = parse_xml(os.path.join(directory, filename))
            
            # Append DataFrame to the specified sheet in the Excel file
            for i in range(df.shape[0]):
                for column_name in df.columns:
                    writer.sheets[sheet_name].cell(row=last_non_empty_row + i + 1, column=column_numbers[column_name], value=df.at[i, column_name])
        
            # Update last_non_empty_row
            last_non_empty_row += df.shape[0]
        except Exception as e:
            print(f"An error occurred while processing the file {filename}: {e}")
            continue

# Save the changes in the Excel file
try:
    writer.save()
except Exception as e:
    print(f"An error occurred while saving the Excel file: {e}")